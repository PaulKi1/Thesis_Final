from __future__ import annotations
import argparse
import re
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook as NewWorkbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from thesis.config import DATABENTO_RAW_DIR, ERIS_SOFR_DIR, INTERIM_DIR

INPUT_FOLDER = DATABENTO_RAW_DIR / "excel"
OUTPUT_FOLDER = INTERIM_DIR
DEFAULT_CURVE_PATH = ERIS_SOFR_DIR / "sofr_discount_curve_lookup.xlsx"
DEFAULT_CURVE_CSV_PATH = ERIS_SOFR_DIR / "sofr_discount_curve_lookup.csv"

DEFAULT_OUTPUT_SHEET = "pre_deamericanization_cleaned"
MAX_EXCEL_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1  # reserve one row for headers

INPUT_EXTENSIONS = {".xlsx", ".xlsm"}

REQUIRED_QUOTE_COLUMNS = [
    "trade_date",
    "symbol",
    "expiration",
    "instrument_class",
    "strike_price",
    "mid_px",
    "staleness_seconds",
    "bid_px_00",
]

REQUIRED_CURVE_COLUMNS = [
    "trade_date",
    "curve_date",
    "discount_factor",
]

TICKER_BRANCH = {
    "NVDA": "underlying",
    "TSLA": "underlying",
    "NVDL": "letf",
    "NVDD": "letf",
    "NVDQ": "letf",
    "TSLR": "letf",
    "TSLS": "letf",
    "TSLT": "letf",
}

BETA_BY_PRODUCT = {
    "NVDA": 1.0,
    "NVDL": 2.0,
    "TSLA": 1.0,
    "TSLR": 2.0,
    "TSLT": 2.0
}

PAIR_GROUP_COLUMNS = ["trade_date", "expiration", "strike_price"]
LOG_MONEYNESS_COLUMN = "Log Moneyness"
LOG_MONEYNESS_DROP_STAT_KEY = "dropped_abs_log_moneyness_gt_abs_max"
MAX_STALENESS_SECONDS = 60
MID_PRICE_TICK_SIZE = 0.01
MIN_MID_PRICE_TICKS = 1
MIN_TIME_TO_EXPIRATION_DAYS = 14
MAX_TIME_TO_EXPIRATION_DAYS: Optional[int] = 365
INTRINSIC_VALUE_MULTIPLIER = 1.01
LOG_MONEYNESS_ABS_MAX = 0.5


def get_effective_log_moneyness_threshold(underlying: str) -> float:
    """
    Return the effective log moneyness threshold for a given underlying/LETF.

    For underlyings (β=1), returns LOG_MONEYNESS_ABS_MAX.
    For LETFs, returns |β| × LOG_MONEYNESS_ABS_MAX so that after moneyness
    scaling the filtered data covers the same range on the underlying's axis.
    """
    branch = TICKER_BRANCH.get(underlying, "underlying")
    if branch == "underlying":
        return LOG_MONEYNESS_ABS_MAX

    beta = BETA_BY_PRODUCT.get(underlying, 1.0)
    return abs(beta) * LOG_MONEYNESS_ABS_MAX


def _normalize_number_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        return str(value)

    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none", "nat"}:
        return None

    text = text.replace("\u00a0", "").replace(" ", "")

    has_comma = "," in text
    has_dot = "." in text

    if has_comma and has_dot:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif has_comma and not has_dot:
        text = text.replace(",", ".")

    return text


def parse_numeric_series(series: pd.Series) -> pd.Series:
    normalized = series.map(_normalize_number_text)
    return pd.to_numeric(normalized, errors="coerce")


_DOT_DATE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
_DOT_DATE_TIME = re.compile(r"^\d{2}\.\d{2}\.\d{4}[ T].*$")
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ISO_DATE_TIME = re.compile(r"^\d{4}-\d{2}-\d{2}[ T].*$")


def parse_date_series(series: pd.Series) -> pd.Series:
    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")

    non_null_mask = series.notna()
    if not non_null_mask.any():
        return parsed

    datetime_mask = series.map(lambda x: isinstance(x, (pd.Timestamp, np.datetime64)))
    if datetime_mask.any():
        parsed.loc[datetime_mask] = pd.to_datetime(series.loc[datetime_mask], errors="coerce")

    text_series = series.loc[non_null_mask & ~datetime_mask].astype(str).str.strip()

    dot_mask = text_series.str.match(_DOT_DATE) | text_series.str.match(_DOT_DATE_TIME)
    if dot_mask.any():
        dot_idx = text_series.index[dot_mask]
        parsed.loc[dot_idx] = pd.to_datetime(text_series.loc[dot_idx], format="%d.%m.%Y", errors="coerce")
        remaining_dot = parsed.loc[dot_idx].isna()
        if remaining_dot.any():
            parsed.loc[dot_idx[remaining_dot]] = pd.to_datetime(
                text_series.loc[dot_idx[remaining_dot]],
                errors="coerce",
                dayfirst=True,
            )

    iso_mask = text_series.str.match(_ISO_DATE) | text_series.str.match(_ISO_DATE_TIME)
    if iso_mask.any():
        iso_idx = text_series.index[iso_mask]
        parsed.loc[iso_idx] = pd.to_datetime(text_series.loc[iso_idx], errors="coerce")

    remaining_mask = parsed.loc[text_series.index].isna()
    if remaining_mask.any():
        remaining_idx = text_series.index[remaining_mask]
        parsed.loc[remaining_idx] = pd.to_datetime(
            text_series.loc[remaining_idx],
            errors="coerce",
            dayfirst=True,
        )

    return parsed


def require_columns(df: pd.DataFrame, required_columns: Iterable[str], dataset_name: str) -> None:
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(f"{dataset_name} is missing required columns: {missing}")


def compute_log_moneyness_series(strike_price: pd.Series, underlying_price: pd.Series) -> pd.Series:
    valid_inputs = strike_price.gt(0) & underlying_price.gt(0)
    return pd.Series(
        np.where(valid_inputs, np.log(strike_price / underlying_price), np.nan),
        index=strike_price.index,
        dtype="float64",
    )


def drop_large_log_moneyness_rows(
        df: pd.DataFrame, log_moneyness_threshold: float = LOG_MONEYNESS_ABS_MAX
) -> tuple[pd.DataFrame, int, float]:
    """
    Drop rows where |Log Moneyness| exceeds the given threshold.

    Returns:
        tuple: (filtered_df, dropped_count, threshold_used)
    """
    if df.empty:
        return df.copy(), 0, log_moneyness_threshold

    require_columns(df, ["strike_price", "underlying_price"], "Log-moneyness filter input")

    filtered = df.copy()
    filtered[LOG_MONEYNESS_COLUMN] = compute_log_moneyness_series(
        filtered["strike_price"],
        filtered["underlying_price"],
    )
    outside_band = filtered[LOG_MONEYNESS_COLUMN].abs().gt(log_moneyness_threshold).fillna(False)
    return filtered.loc[~outside_band].copy(), int(outside_band.sum()), log_moneyness_threshold


def drop_unmatched_option_sides(df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    if df.empty:
        return df.copy(), 0, 0

    normalized_side = df["instrument_class"].astype("string").str.upper().str.strip()
    grouped = (
        df.assign(_instrument_class_norm=normalized_side)
        .groupby(PAIR_GROUP_COLUMNS)["_instrument_class_norm"]
    )
    has_call = grouped.transform(lambda side: side.eq("C").any())
    has_put = grouped.transform(lambda side: side.eq("P").any())
    keep_mask = has_call & has_put

    dropped_rows = int((~keep_mask).sum())
    dropped_groups = int(df.loc[~keep_mask, PAIR_GROUP_COLUMNS].drop_duplicates().shape[0])
    return df.loc[keep_mask].copy(), dropped_rows, dropped_groups


def extract_underlying_from_symbol(symbol_series: pd.Series) -> pd.Series:
    extracted = symbol_series.astype(str).str.extract(r"^\s*([^\s]+)", expand=False)
    return extracted.fillna("").str.strip()


def prepare_curve_lookup(curve_df: pd.DataFrame) -> pd.DataFrame:
    curve = curve_df.copy()
    curve.columns = [str(column).strip() for column in curve.columns]
    require_columns(curve, REQUIRED_CURVE_COLUMNS, "Discount curve file")

    curve["trade_date"] = parse_date_series(curve["trade_date"]).dt.normalize()
    curve["curve_date"] = parse_date_series(curve["curve_date"]).dt.normalize()
    curve["discount_factor"] = parse_numeric_series(curve["discount_factor"])

    invalid_key_rows = curve["trade_date"].isna() | curve["curve_date"].isna()
    if invalid_key_rows.any():
        raise ValueError(
            f"Discount curve file has {int(invalid_key_rows.sum())} row(s) with unreadable trade_date or curve_date."
        )

    if curve["discount_factor"].isna().any():
        raise ValueError(
            f"Discount curve file has {int(curve['discount_factor'].isna().sum())} row(s) with unreadable discount_factor."
        )

    duplicate_mask = curve.duplicated(subset=["trade_date", "curve_date"], keep=False)
    if duplicate_mask.any():
        duplicate_rows = curve.loc[duplicate_mask, ["trade_date", "curve_date"]].drop_duplicates()
        raise ValueError(
            "Discount curve lookup contains duplicate (trade_date, curve_date) keys. "
            f"Examples: {duplicate_rows.head(5).to_dict(orient='records')}"
        )

    return curve[["trade_date", "curve_date", "discount_factor"]].copy()


def clean_quotes(quotes_df: pd.DataFrame, curve_lookup: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int | float]]:
    df = quotes_df.copy()
    df.columns = [str(column).strip() for column in df.columns]
    require_columns(df, REQUIRED_QUOTE_COLUMNS, "Input quote sheet")

    price_column = "underlying_price_unadj" if "underlying_price_unadj" in df.columns else "underlying_price"
    if price_column not in df.columns:
        raise ValueError("Input quote sheet must contain 'underlying_price_unadj' or 'underlying_price'.")

    df["trade_date"] = parse_date_series(df["trade_date"]).dt.normalize()
    df["expiration"] = parse_date_series(df["expiration"]).dt.normalize()
    df["mid_px"] = parse_numeric_series(df["mid_px"])
    df["strike_price"] = parse_numeric_series(df["strike_price"])
    df["staleness_seconds"] = parse_numeric_series(df["staleness_seconds"])
    df["bid_px_00"] = parse_numeric_series(df["bid_px_00"])
    df["underlying_price"] = parse_numeric_series(df[price_column])
    df["instrument_class"] = df["instrument_class"].astype(str).str.strip().str.upper()
    df["symbol"] = df["symbol"].astype(str).str.strip()
    df["underlying"] = extract_underlying_from_symbol(df["symbol"])

    stats: dict[str, int | float] = {"rows_in": int(len(df))}

    missing_mid_mask = df["mid_px"].isna()
    stats["dropped_missing_mid_px"] = int(missing_mid_mask.sum())
    df = df.loc[~missing_mid_mask].copy()

    min_mid_price = MIN_MID_PRICE_TICKS * MID_PRICE_TICK_SIZE
    low_mid_price_mask = df["mid_px"] < min_mid_price
    stats["dropped_mid_px_below_min_ticks"] = int(low_mid_price_mask.sum())
    df = df.loc[~low_mid_price_mask].copy()

    stale_mask = df["staleness_seconds"].notna() & (df["staleness_seconds"] > MAX_STALENESS_SECONDS)
    stats["dropped_stale_quotes"] = int(stale_mask.sum())
    df = df.loc[~stale_mask].copy()

    zero_bid_mask = df["bid_px_00"].notna() & (df["bid_px_00"] == 0)
    stats["dropped_zero_bid_px_00"] = int(zero_bid_mask.sum())
    df = df.loc[~zero_bid_mask].copy()

    unreadable_required_mask = (
            df["trade_date"].isna()
            | df["expiration"].isna()
            | df["strike_price"].isna()
            | df["underlying_price"].isna()
            | ~df["instrument_class"].isin(["C", "P"])
    )
    stats["dropped_invalid_required_fields"] = int(unreadable_required_mask.sum())
    df = df.loc[~unreadable_required_mask].copy()

    call_mask = df["instrument_class"] == "C"
    put_mask = df["instrument_class"] == "P"

    intrinsic = np.where(call_mask, np.maximum(df["underlying_price"] - df["strike_price"], 0.0),
                         np.where(put_mask, np.maximum(df["strike_price"] - df["underlying_price"], 0.0), np.nan),
                         )
    df["intrinsic_value"] = intrinsic

    near_intrinsic_mask = ~(df["mid_px"] > INTRINSIC_VALUE_MULTIPLIER * df["intrinsic_value"])
    stats["dropped_not_1pct_above_intrinsic"] = int(near_intrinsic_mask.sum())
    df = df.loc[~near_intrinsic_mask].copy()

    df["time_to_expiration"] = (df["expiration"] - df["trade_date"]).dt.days
    invalid_tte_mask = df["time_to_expiration"].isna()
    stats["dropped_invalid_time_to_expiration"] = int(invalid_tte_mask.sum())
    df = df.loc[~invalid_tte_mask].copy()
    short_tte_mask = df["time_to_expiration"] <= MIN_TIME_TO_EXPIRATION_DAYS
    stats["dropped_short_time_to_expiration"] = int(short_tte_mask.sum())
    df = df.loc[~short_tte_mask].copy()
    if MAX_TIME_TO_EXPIRATION_DAYS is None:
        stats["dropped_long_time_to_expiration"] = 0
    else:
        long_tte_mask = df["time_to_expiration"] > MAX_TIME_TO_EXPIRATION_DAYS
        stats["dropped_long_time_to_expiration"] = int(long_tte_mask.sum())
        df = df.loc[~long_tte_mask].copy()

    # Determine effective log moneyness threshold based on underlying/LETF
    # For LETFs, threshold is scaled by |β| for proper moneyness alignment
    unique_underlyings = df["underlying"].dropna().unique()
    if len(unique_underlyings) == 1:
        ticker = unique_underlyings[0]
    elif len(unique_underlyings) > 1:
        # Multiple underlyings in one sheet - use the most common one
        ticker = df["underlying"].mode().iloc[0] if not df["underlying"].mode().empty else ""
    else:
        ticker = ""

    effective_lm_threshold = get_effective_log_moneyness_threshold(ticker)
    df, stats[LOG_MONEYNESS_DROP_STAT_KEY], threshold_used = drop_large_log_moneyness_rows(
        df, log_moneyness_threshold=effective_lm_threshold
    )
    stats["log_moneyness_threshold_used"] = threshold_used

    merged = df.merge(
        curve_lookup,
        how="left",
        left_on=["trade_date", "expiration"],
        right_on=["trade_date", "curve_date"],
        validate="m:1",
    )

    missing_df_mask = merged["discount_factor"].isna()
    stats["dropped_missing_discount_factor"] = int(missing_df_mask.sum())
    if missing_df_mask.any():
        missing_examples = merged.loc[missing_df_mask, ["trade_date", "expiration"]].drop_duplicates().head(10)
        raise ValueError(
            "Missing discount_factor for some (trade_date, expiration) pairs. "
            f"Examples: {missing_examples.to_dict(orient='records')}"
        )

    merged, stats["dropped_unmatched_option_side_rows"], stats["dropped_unmatched_option_side_groups"] = (
        drop_unmatched_option_sides(merged)
    )

    output = merged[
        [
            "underlying",
            "trade_date",
            "symbol",
            "expiration",
            "instrument_class",
            "strike_price",
            "mid_px",
            "underlying_price",
            "time_to_expiration",
            "discount_factor",
        ]
    ].copy()

    # No sorting here — global sort happens in process_single_file
    output = output.reset_index(drop=True)
    stats["rows_out"] = int(len(output))
    return output, stats


def build_output_path(input_workbook: Path, output_folder: Path) -> Path:
    return output_folder / f"{input_workbook.stem}_with_cleaned_sheet.xlsx"


def discover_relevant_input_sheets(input_workbook: Path, requested_sheet: Optional[str]) -> tuple[list[str], list[str]]:
    if requested_sheet:
        preview = pd.read_excel(input_workbook, sheet_name=requested_sheet, nrows=0, dtype=object)
        preview.columns = [str(column).strip() for column in preview.columns]
        require_columns(preview, REQUIRED_QUOTE_COLUMNS, f"Requested input sheet '{requested_sheet}'")
        if "underlying_price_unadj" not in preview.columns and "underlying_price" not in preview.columns:
            raise ValueError(
                f"Requested input sheet '{requested_sheet}' must contain 'underlying_price_unadj' or 'underlying_price'."
            )
        return [requested_sheet], []

    workbook = load_workbook(input_workbook, read_only=True)
    relevant: list[str] = []
    skipped: list[str] = []

    for sheet_name in workbook.sheetnames:
        preview = pd.read_excel(input_workbook, sheet_name=sheet_name, nrows=0, dtype=object)
        preview.columns = [str(column).strip() for column in preview.columns]
        required_present = all(column in preview.columns for column in REQUIRED_QUOTE_COLUMNS)
        has_price = "underlying_price_unadj" in preview.columns or "underlying_price" in preview.columns
        if required_present and has_price:
            relevant.append(sheet_name)
        else:
            skipped.append(sheet_name)

    return relevant, skipped


def _safe_sheet_name(base_name: str, suffix_number: Optional[int] = None) -> str:
    if suffix_number is None:
        return base_name[:31]

    suffix = f"_{suffix_number}"
    max_base_length = 31 - len(suffix)
    return f"{base_name[:max_base_length]}{suffix}"


class CleanedSheetWriter:
    OUTPUT_COLUMNS = [
        "underlying",
        "trade_date",
        "symbol",
        "expiration",
        "instrument_class",
        "strike_price",
        "mid_px",
        "underlying_price",
        "time_to_expiration",
        "discount_factor",
    ]

    def __init__(self, workbook, base_sheet_name: str):
        self.workbook = workbook
        self.base_sheet_name = base_sheet_name
        self.sheet_counter = 0
        self.current_sheet = None
        self.current_sheet_row = 0  # data rows only
        self.total_rows_written = 0
        self.sheet_names: list[str] = []

        self.header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.center_alignment = Alignment(horizontal="center")

        self._remove_existing_output_sheets()

    def _remove_existing_output_sheets(self) -> None:
        to_delete = []
        for sheet_name in self.workbook.sheetnames:
            if sheet_name == self.base_sheet_name or sheet_name.startswith(f"{self.base_sheet_name}_"):
                to_delete.append(sheet_name)
        for sheet_name in to_delete:
            del self.workbook[sheet_name]

    def _ensure_sheet(self) -> None:
        if self.current_sheet is None:
            self.sheet_counter = 1
            self.current_sheet = self.workbook.create_sheet(title=_safe_sheet_name(self.base_sheet_name))
            self.current_sheet.freeze_panes = "A2"
            self._write_header(self.current_sheet)
            self.current_sheet_row = 0
            self.sheet_names = [self.current_sheet.title]
            return

        if self.current_sheet_row >= MAX_DATA_ROWS_PER_SHEET:
            if self.sheet_counter == 1 and self.current_sheet.title == _safe_sheet_name(self.base_sheet_name):
                self.current_sheet.title = _safe_sheet_name(self.base_sheet_name, 1)
                self.sheet_names[0] = self.current_sheet.title

            self.sheet_counter += 1
            next_title = _safe_sheet_name(self.base_sheet_name, self.sheet_counter)
            self.current_sheet = self.workbook.create_sheet(title=next_title)
            self.current_sheet.freeze_panes = "A2"
            self._write_header(self.current_sheet)
            self.current_sheet_row = 0
            self.sheet_names.append(self.current_sheet.title)

    def _write_header(self, worksheet) -> None:
        for column_index, header in enumerate(self.OUTPUT_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(len(header) + 2, 12)
        last_column_letter = get_column_letter(len(self.OUTPUT_COLUMNS))
        worksheet.auto_filter.ref = f"A1:{last_column_letter}1"

    @staticmethod
    def _python_value(column_name: str, value: object) -> object:
        if pd.isna(value):
            return None
        if column_name in {"trade_date", "expiration"}:
            return pd.Timestamp(value).to_pydatetime()
        if column_name == "time_to_expiration":
            return int(value)
        return value

    @staticmethod
    def _apply_number_format(cell, column_name: str) -> None:
        if column_name in {"trade_date", "expiration"}:
            cell.number_format = "DD.MM.YYYY"
        elif column_name in {"strike_price"}:
            cell.number_format = "0"
        elif column_name in {"mid_px", "underlying_price"}:
            cell.number_format = "0.000"
        elif column_name == "time_to_expiration":
            cell.number_format = "0"
        elif column_name == "discount_factor":
            cell.number_format = "0.000000"

    @staticmethod
    def _update_width(worksheet, column_index: int, value: object) -> None:
        column_letter = get_column_letter(column_index)
        width = worksheet.column_dimensions[column_letter].width or 12
        text = "" if value is None else str(value)
        worksheet.column_dimensions[column_letter].width = min(max(width, len(text) + 2), 40)

    def append_dataframe(self, cleaned_df: pd.DataFrame) -> None:
        if cleaned_df.empty:
            return

        for row in cleaned_df[self.OUTPUT_COLUMNS].itertuples(index=False, name=None):
            self._ensure_sheet()
            excel_row = self.current_sheet_row + 2

            for column_index, (column_name, raw_value) in enumerate(zip(self.OUTPUT_COLUMNS, row), start=1):
                value = self._python_value(column_name, raw_value)
                cell = self.current_sheet.cell(row=excel_row, column=column_index, value=value)
                self._apply_number_format(cell, column_name)
                self._update_width(self.current_sheet, column_index, value)

            self.current_sheet_row += 1
            self.total_rows_written += 1
            last_column_letter = get_column_letter(len(self.OUTPUT_COLUMNS))
            self.current_sheet.auto_filter.ref = f"A1:{last_column_letter}{self.current_sheet_row + 1}"

    def finalize(self) -> None:
        if self.sheet_counter == 0:
            worksheet = self.workbook.create_sheet(title=_safe_sheet_name(self.base_sheet_name))
            worksheet.freeze_panes = "A2"
            self._write_header(worksheet)
            self.sheet_names = [worksheet.title]


def discover_input_files(input_folder: Path) -> list[Path]:
    """Find all .xlsx and .xlsm files in the input folder (non-recursive)."""
    files = sorted(
        f for f in input_folder.iterdir()
        if f.is_file() and f.suffix.lower() in INPUT_EXTENSIONS
    )
    return files


def process_single_file(
        input_workbook: Path,
        curve_lookup: pd.DataFrame,
        output_folder: Path,
        input_sheet: Optional[str],
        output_sheet: str,
) -> None:
    """Process one input workbook: clean its sheets and write the Output reverse Loop file."""
    print(f"\n{'=' * 70}")
    print(f"Processing: {input_workbook.name}")
    print(f"{'=' * 70}")

    relevant_sheets, skipped_sheets = discover_relevant_input_sheets(input_workbook, input_sheet)
    if not relevant_sheets:
        print(f"  SKIPPED – no relevant quote sheets found in {input_workbook.name}")
        return

    output_workbook = build_output_path(input_workbook, output_folder)

    # Create a fresh workbook (only the cleaned sheet will exist)
    workbook = NewWorkbook()
    workbook.remove(workbook.active)  # drop the default empty sheet

    writer = CleanedSheetWriter(workbook, output_sheet)

    total_stats = {
        "sheets_processed": 0,
        "rows_in": 0,
        "dropped_missing_mid_px": 0,
        "dropped_mid_px_below_min_ticks": 0,
        "dropped_stale_quotes": 0,
        "dropped_zero_bid_px_00": 0,
        "dropped_invalid_required_fields": 0,
        "dropped_not_1pct_above_intrinsic": 0,
        "dropped_invalid_time_to_expiration": 0,
        "dropped_short_time_to_expiration": 0,
        "dropped_long_time_to_expiration": 0,
        LOG_MONEYNESS_DROP_STAT_KEY: 0,
        "dropped_missing_discount_factor": 0,
        "dropped_unmatched_option_side_rows": 0,
        "dropped_unmatched_option_side_groups": 0,
        "rows_out": 0,
        "unique_trade_dates": 0,
    }

    per_sheet_stats: list[tuple[str, dict[str, int]]] = []

    # ── Collect all cleaned data ──
    all_cleaned: list[pd.DataFrame] = []

    for sheet_name in relevant_sheets:
        quotes_df = pd.read_excel(input_workbook, sheet_name=sheet_name, dtype=object)
        cleaned_df, stats = clean_quotes(quotes_df, curve_lookup)

        total_stats["sheets_processed"] += 1
        for key, value in stats.items():
            total_stats[key] = total_stats.get(key, 0) + int(value)
        per_sheet_stats.append((sheet_name, stats))

        if not cleaned_df.empty:
            all_cleaned.append(cleaned_df)

    # ── Merge, sort globally, then write once ──
    if all_cleaned:
        combined = pd.concat(all_cleaned, ignore_index=True)
        combined = combined.sort_values(
            ["trade_date", "underlying", "expiration", "symbol"]
        ).reset_index(drop=True)

        total_stats["rows_out"] = int(len(combined))
        total_stats["unique_trade_dates"] = int(combined["trade_date"].nunique())

        writer.append_dataframe(combined)

    writer.finalize()
    workbook.save(output_workbook)

    print(f"  Output reverse Loop workbook: {output_workbook}")
    print(f"  Relevant input sheets processed: {relevant_sheets}")
    if skipped_sheets:
        print(f"  Skipped non-quote sheets: {skipped_sheets}")
    print(f"  Created Output reverse Loop sheets: {writer.sheet_names}")
    for key, value in total_stats.items():
        print(f"  total_{key}: {value}")

    if len(per_sheet_stats) > 1:
        print("  Per-sheet stats:")
        for sheet_name, stats in per_sheet_stats:
            stats_text = ", ".join(f"{key}={value}" for key, value in stats.items())
            print(f"    - {sheet_name}: {stats_text}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Standalone pre-de-Americanization cleaning script. "
            "Processes all .xlsx/.xlsm files in the input folder (or a single file). "
            "For each file it scans one specified input sheet or all relevant sheets, "
            "drops rows with missing mid_px, "
            f"mid_px below {MIN_MID_PRICE_TICKS * MID_PRICE_TICK_SIZE:.2f} "
            f"({MIN_MID_PRICE_TICKS} tick(s) at {MID_PRICE_TICK_SIZE:.2f} each), "
            "drops stale quotes with "
            f"staleness_seconds > {MAX_STALENESS_SECONDS}, "
            "drops rows where bid_px_00 equals zero, "
            "drops quotes whose midpoint is not more than 1% above American intrinsic value, "
            "drops rows with time to expiration in calendar days above "
            "MAX_TIME_TO_EXPIRATION_DAYS when configured, "
            f"drops rows with |Log Moneyness| > {LOG_MONEYNESS_ABS_MAX:.1f} after computing "
            "Log Moneyness internally as ln(strike_price / underlying_price), "
            "drops strike slices that no longer contain both a call and a put, "
            "computes time_to_expiration in calendar days internally, maps discount_factor from an exact "
            "(trade_date, expiration) lookup, and writes the cleaned Output reverse Loop into as few Output reverse Loop "
            "sheets as possible without exceeding Excel's row limit."
        )
    )
    parser.add_argument(
        "--input-folder",
        default=None,
        help=(
            "Folder containing the raw option workbooks (.xlsx/.xlsm). "
            "Defaults to INPUT_FOLDER constant at the top of the script."
        ),
    )
    parser.add_argument(
        "--input-workbook",
        default=None,
        help=(
            "Path to a single raw option workbook (.xlsx or .xlsm). "
            "If provided, only this file is processed instead of scanning the input folder."
        ),
    )
    parser.add_argument(
        "--input-sheet",
        default=None,
        help="Optional single input sheet name. If omitted, the script scans all relevant sheets in each workbook.",
    )
    parser.add_argument(
        "--curve-workbook",
        default=DEFAULT_CURVE_PATH,
        help="Path to the discount curve lookup workbook.",
    )
    parser.add_argument(
        "--curve-sheet",
        default=None,
        help="Sheet name in the discount curve workbook. Defaults to the first sheet if omitted.",
    )
    parser.add_argument(
        "--output-folder",
        default=None,
        help=(
            "Folder where the Output reverse Loop workbooks will be saved. "
            "Defaults to OUTPUT_FOLDER constant at the top of the script."
        ),
    )
    parser.add_argument(
        "--output-sheet",
        default=DEFAULT_OUTPUT_SHEET,
        help=(
            "Base name of the cleaned Output reverse Loop sheet. If multiple cleaned sheets are needed, "
            "the script creates '<base>_1', '<base>_2', and so on."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_folder = Path(args.input_folder) if args.input_folder else Path(INPUT_FOLDER)
    output_folder = Path(args.output_folder) if args.output_folder else Path(OUTPUT_FOLDER)
    curve_workbook = Path(args.curve_workbook)

    if not curve_workbook.exists():
        if curve_workbook == DEFAULT_CURVE_PATH and DEFAULT_CURVE_CSV_PATH.exists():
            curve_workbook = DEFAULT_CURVE_CSV_PATH
        else:
            raise FileNotFoundError(f"Discount curve workbook not found: {curve_workbook}")

    # Determine which input files to process
    if args.input_workbook:
        # Single-file mode (backward compatible)
        single_path = Path(args.input_workbook)
        if not single_path.exists():
            raise FileNotFoundError(f"Input workbook not found: {single_path}")
        input_files = [single_path]
    else:
        # Folder mode
        if not input_folder.exists():
            raise FileNotFoundError(f"Input folder not found: {input_folder}")
        input_files = discover_input_files(input_folder)
        if not input_files:
            raise FileNotFoundError(
                f"No .xlsx or .xlsm files found in: {input_folder}"
            )

    output_folder.mkdir(parents=True, exist_ok=True)

    # Load the discount curve once (shared across all files)
    print(f"Loading discount curve from: {curve_workbook}")
    if curve_workbook.suffix.lower() == ".csv":
        curve_df = pd.read_csv(curve_workbook, dtype=object)
    else:
        curve_df = pd.read_excel(curve_workbook, sheet_name=args.curve_sheet or 0, dtype=object)
    curve_lookup = prepare_curve_lookup(curve_df)
    print(f"Discount curve loaded: {len(curve_lookup)} rows")

    print(f"\nInput files to process: {len(input_files)}")
    print(f"Output reverse Loop folder: {output_folder}")

    succeeded = 0
    failed = 0
    failed_files: list[tuple[str, str]] = []

    for input_file in input_files:
        try:
            process_single_file(
                input_workbook=input_file,
                curve_lookup=curve_lookup,
                output_folder=output_folder,
                input_sheet=args.input_sheet,
                output_sheet=args.output_sheet,
            )
            succeeded += 1
        except Exception as exc:
            failed += 1
            failed_files.append((input_file.name, str(exc)))
            print(f"\n  ERROR processing {input_file.name}: {exc}")

    # Final summary
    print(f"\n{'=' * 70}")
    print(f"BATCH COMPLETE: {succeeded} succeeded, {failed} failed out of {len(input_files)} files")
    print(f"{'=' * 70}")
    if failed_files:
        print("Failed files:")
        for filename, error_msg in failed_files:
            print(f"  - {filename}: {error_msg}")


if __name__ == "__main__":
    main()
